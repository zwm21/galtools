# -*- coding: utf-8 -*-
"""vndb_voiced 的离线测试。

整套测试零联网：api 模块只有 _open / _sleep 两个注入点，全部测试都打这两处。
_sleep 顺带推进一个假时钟——限流是按 time.monotonic 做的滑动窗口，假睡眠不推
时钟的话窗口永远滚不过去，会死循环。
"""
import io
import json
import os
import re
import urllib.error
from email.message import Message

import pytest

from galtools.core.context import Cancelled, RunContext
from galtools.tools import vndb_voiced as tool
from galtools.tools.vndb_voiced import api, cli, fetch, xlsx
from galtools.tools.vndb_voiced.model import (
    Common, Credit, Resolution, Staff, StaffCredits, label, released_sort_key,
    url_for,
)


# ---------------- 假传输 ----------------
class FakeClock:
    def __init__(self):
        self.now = 1000.0

    def monotonic(self):
        return self.now


class FakeResponse:
    """够用的 urlopen 返回值：分块 read + close。payload 给 bytes 就原样发。"""

    def __init__(self, payload):
        self.body = (payload if isinstance(payload, bytes)
                     else json.dumps(payload).encode('utf-8'))
        self.closed = False

    def read(self, size=-1):
        if size is None or size < 0:
            out, self.body = self.body, b''
            return out
        out, self.body = self.body[:size], self.body[size:]
        return out

    def close(self):
        self.closed = True


def http_error(code, body=b'', retry_after=None):
    headers = Message()
    if retry_after is not None:
        headers['Retry-After'] = retry_after
    return urllib.error.HTTPError('u', code, 'msg', headers, io.BytesIO(body))


def page(results, more=False, count=None):
    out = {'results': list(results), 'more': more}
    if count is not None:
        out['count'] = count
    return out


class FakeApi:
    """handler(path, body, nth) 返回 payload 或要抛的异常。"""

    def __init__(self, handler):
        self.handler = handler
        self.calls = []
        self.sleeps = []
        self.clock = FakeClock()

    def install(self, monkeypatch):
        monkeypatch.setattr(api, '_open', self.open)
        monkeypatch.setattr(api, '_sleep', self.sleep)
        monkeypatch.setattr(api, 'time', self.clock)
        return self

    def open(self, req, timeout):
        path = req.full_url[len(api.API_BASE):]
        body = json.loads(req.data.decode('utf-8'))
        self.calls.append((path, body))
        out = self.handler(path, body, len(self.calls))
        if isinstance(out, BaseException):
            raise out
        return FakeResponse(out)

    def sleep(self, seconds):
        self.sleeps.append(seconds)
        self.clock.now += seconds

    def paths(self):
        return [p for p, _ in self.calls]


# ---------------- 纯函数 ----------------
def test_normalize_ignores_space_and_case():
    assert fetch.normalize('水橋 かおり') == fetch.normalize('水橋かおり')
    assert fetch.normalize('Ono Ryouko') == fetch.normalize('ONORYOUKO')


def test_parse_targets_keeps_names_with_spaces():
    # 罗马字人名自带空格，不能拿空格当分隔符。
    assert fetch.parse_targets('Ono Ryouko, 水橋 かおり') == ['Ono Ryouko', '水橋 かおり']


def test_parse_targets_splits_ids_by_space():
    assert fetch.parse_targets('s367 s131\ns2') == ['s367', 's131', 's2']


def test_parse_targets_dedupes_by_normalized_form():
    assert fetch.parse_targets('水橋 かおり、水橋かおり') == ['水橋 かおり']
    assert fetch.parse_targets(' ,, \n') == []


@pytest.mark.parametrize('raw, expected', [
    ('s367', ('id', 's367')),
    ('367', ('id', 's367')),
    ('S367', ('id', 's367')),
    ('https://vndb.org/s367', ('id', 's367')),
    ('vndb.org/S367/chars', ('id', 's367')),
    ('Ono Ryouko', ('name', 'Ono Ryouko')),
])
def test_classify_good(raw, expected):
    assert fetch.classify(raw) == expected


@pytest.mark.parametrize('raw, hint', [
    ('', '为空'),
    ('a', '太短'),
    ('v17', '作品'),
    ('c123', '角色'),
    ('https://vndb.org/v17', '作品'),
    ('http://example.com/x', 'staff id'),
])
def test_classify_bad(raw, hint):
    kind, reason = fetch.classify(raw)
    assert kind == 'bad' and hint in reason


def test_released_sort_key_puts_year_only_before_full_date():
    # 旧 compare 脚本把 '1995' 折成 '9999' 排到 TBA 后面，这里修掉。
    assert released_sort_key('1995') < released_sort_key('1995-01-31')
    assert released_sort_key('1995-01') < released_sort_key('1995-01-31')
    assert released_sort_key('1995-12-31') < released_sort_key('1996')
    assert released_sort_key('9999-99-99') == released_sort_key(None)
    assert released_sort_key('1995-01-31') < released_sort_key('')
    assert released_sort_key('TBA') == released_sort_key(None)


def test_label_and_url():
    assert label('Ono Ryouko', '小野 涼子', 's367') == '小野 涼子 (Ono Ryouko) s367'
    assert label('Aoi Kaoru', 'Aoi Kaoru', 's1') == 'Aoi Kaoru s1'
    assert label('Aoi Kaoru', '', '') == 'Aoi Kaoru'
    assert url_for('v17') == 'https://vndb.org/v17'
    assert url_for('') is None


def test_orig_falls_back_to_the_latin_text():
    # 原名本就是拉丁字母的作品/角色，接口不给 alttitle 与 original，
    # 「原名」三列不能因此留空。
    full = Credit(vid='v1', title='Aoi', title_ja='蒼', cast='Chara',
                  cast_ja='キャラ', alias='Al', alias_ja='アル')
    assert (full.title_orig, full.cast_orig, full.alias_orig) == ('蒼', 'キャラ',
                                                                 'アル')
    bare = Credit(vid='v2', title='Aoi', cast='Chara', alias='Al')
    assert (bare.title_orig, bare.cast_orig, bare.alias_orig) == ('Aoi', 'Chara',
                                                                  'Al')
    assert Common(vid='v1', title='Aoi').title_orig == 'Aoi'
    assert Common(vid='v1', title='Aoi', title_ja='蒼').title_orig == '蒼'


# ---------------- api ----------------
def test_paged_walks_pages_and_counts_once(monkeypatch):
    def handler(path, body, nth):
        assert body['results'] == api.PAGE_SIZE
        if body['page'] == 1:
            assert body['count'] is True
            return page([{'id': 'v1'}], more=True, count=3)
        assert 'count' not in body
        return page([{'id': 'v2'}, {'id': 'v3'}])

    fake = FakeApi(handler).install(monkeypatch)
    client = api.Client()
    seen = []
    items = list(client.paged('vn', {'fields': 'id'}, on_count=seen.append))
    assert [i['id'] for i in items] == ['v1', 'v2', 'v3']
    assert seen == [3] and client.requests == 2 and fake.paths() == ['vn', 'vn']


def test_count_asks_for_one_row(monkeypatch):
    fake = FakeApi(lambda p, b, n: page([{'id': 'v1'}], count=268)).install(monkeypatch)
    assert api.Client().count('vn', ['id', '=', 'v1']) == 268
    assert fake.calls[0][1] == {'filters': ['id', '=', 'v1'], 'fields': 'id',
                                'results': 1, 'count': True}


def test_server_error_is_retried(monkeypatch):
    fake = FakeApi(lambda p, b, n: http_error(503) if n == 1
                   else page([{'id': 'v1'}])).install(monkeypatch)
    assert api.Client().post('vn', {})['results'] == [{'id': 'v1'}]
    assert len(fake.calls) == 2
    assert sum(fake.sleeps) == pytest.approx(1.0, abs=0.01)   # 退避第一档


def test_429_honours_retry_after(monkeypatch):
    fake = FakeApi(lambda p, b, n: http_error(429, retry_after='3') if n == 1
                   else page([])).install(monkeypatch)
    api.Client().post('vn', {})
    assert sum(fake.sleeps) == pytest.approx(3.0, abs=0.01)
    # 睡眠被切片，每片之间才有机会响应取消。
    assert max(fake.sleeps) <= api.SLEEP_SLICE


@pytest.mark.parametrize('value', ['999', 'later', '-1'])
def test_absurd_retry_after_falls_back_to_backoff(monkeypatch, value):
    fake = FakeApi(lambda p, b, n: http_error(429, retry_after=value) if n == 1
                   else page([])).install(monkeypatch)
    api.Client().post('vn', {})
    assert sum(fake.sleeps) == pytest.approx(1.0, abs=0.01)


def test_400_is_not_retried_and_quotes_body(monkeypatch):
    fake = FakeApi(lambda p, b, n: http_error(400, b'Invalid filter')).install(monkeypatch)
    with pytest.raises(api.ApiError) as err:
        api.Client().post('vn', {})
    assert 'Invalid filter' in str(err.value) and len(fake.calls) == 1


def test_network_error_gives_up_after_max_attempts(monkeypatch):
    fake = FakeApi(lambda p, b, n: urllib.error.URLError('unreachable')).install(monkeypatch)
    with pytest.raises(api.ApiError) as err:
        api.Client().post('vn', {})
    assert len(fake.calls) == api.MAX_ATTEMPTS
    assert 'api.vndb.org' in str(err.value) and '重试' in str(err.value)


def test_garbage_body_is_reported(monkeypatch):
    FakeApi(lambda p, b, n: b'<html>oops</html>').install(monkeypatch)
    with pytest.raises(api.ApiError) as err:
        api.Client().post('vn', {})
    assert 'JSON' in str(err.value)


class CountingCtx(RunContext):
    """第 after 次 check_cancel 起抛 Cancelled，用来钉住取消检查的位置。"""

    def __init__(self, after):
        super().__init__()
        self.after = after
        self.checks = 0
        self.logs = []
        self.progress_calls = []

    def check_cancel(self):
        self.checks += 1
        if self.checks >= self.after:
            raise Cancelled()

    def log(self, msg, level='info'):
        self.logs.append((level, msg))

    def progress(self, done, total, note=''):
        self.progress_calls.append((done, total, note))


def test_cancel_before_request(monkeypatch):
    fake = FakeApi(lambda p, b, n: page([])).install(monkeypatch)
    with pytest.raises(Cancelled):
        api.Client(CountingCtx(1)).post('vn', {})
    assert fake.calls == []          # 一次请求都没发出


def test_cancel_while_reading_body(monkeypatch):
    fake = FakeApi(lambda p, b, n: page([{'id': 'v1'}])).install(monkeypatch)
    with pytest.raises(Cancelled):
        api.Client(CountingCtx(2)).post('vn', {})
    assert len(fake.calls) == 1      # 发出去了，读的时候才取消


def test_throttle_waits_for_window_to_slide(monkeypatch):
    fake = FakeApi(lambda p, b, n: page([])).install(monkeypatch)
    client = api.Client(quota=2)
    for _ in range(3):
        client.post('vn', {})
    assert sum(fake.sleeps) > api.RATE_WINDOW - 1
    assert len(fake.calls) == 3


def test_min_interval_between_requests(monkeypatch):
    fake = FakeApi(lambda p, b, n: page([])).install(monkeypatch)
    client = api.Client()
    client.post('vn', {})
    client.post('vn', {})
    assert sum(fake.sleeps) == pytest.approx(api.MIN_INTERVAL, abs=0.01)


# ---------------- 一个迷你 vndb ----------------
# s1 有两个别名（a1 主名、a2 别名）；v1 里 s1 配了 c1 与 c2，s2 配了 c3；
# c2 还在 v2 里出现过（v2 由 s2 配），所以 v2 会被嵌套 filter 当成 s1 的候选
# 作品带回来——正是必须在客户端二次过滤掉的那一类。
STAFF_ROWS = {
    's1': [{'id': 's1', 'aid': 'a1', 'ismain': True,
            'name': 'Alpha One', 'original': 'アルファ'},
           {'id': 's1', 'aid': 'a2', 'ismain': False,
            'name': 'A. One', 'original': ''}],
    's2': [{'id': 's2', 'aid': 'a9', 'ismain': True,
            'name': 'Beta Two', 'original': 'ベータ'}],
    's3': [{'id': 's3', 'aid': 'a7', 'ismain': True,
            'name': 'Gamma Three', 'original': 'ガンマ'}],
}
CHAR_ROWS = {
    's1': [{'id': 'c1', 'name': 'Chara One', 'original': 'キャラ壱',
            'vns': [{'id': 'v1', 'role': 'main'}]},
           {'id': 'c2', 'name': 'Chara Two', 'original': '',
            'vns': [{'id': 'v1', 'role': 'side'}, {'id': 'v2', 'role': 'appears'}]}],
    's2': [{'id': 'c3', 'name': 'Chara Three', 'original': 'キャラ参',
            'vns': [{'id': 'v1', 'role': 'primary'}, {'id': 'v2', 'role': 'main'}]}],
    's3': [{'id': 'c4', 'name': 'Chara Four', 'original': 'キャラ肆',
            'vns': [{'id': 'v1', 'role': 'side'}, {'id': 'v2', 'role': 'main'}]}],
}
VN_ROWS = {
    'v1': {'id': 'v1', 'title': 'Game One', 'alttitle': 'ゲーム壱',
           'released': '1995',
           'va': [{'staff': {'id': 's1', 'aid': 'a1'},
                   'character': {'id': 'c1'}, 'note': ''},
                  {'staff': {'id': 's1', 'aid': 'a2'},
                   'character': {'id': 'c2'}, 'note': '2024 remake'},
                  {'staff': {'id': 's2', 'aid': 'a9'},
                   'character': {'id': 'c3'}, 'note': ''},
                  {'staff': {'id': 's3', 'aid': 'a7'},
                   'character': {'id': 'c4'}, 'note': ''}]},
    'v2': {'id': 'v2', 'title': 'Game Two', 'alttitle': '', 'released': '1995-01-31',
           'va': [{'staff': {'id': 's2', 'aid': 'a9'},
                   'character': {'id': 'c3'}, 'note': ''},
                  {'staff': {'id': 's3', 'aid': 'a7'},
                   'character': {'id': 'c4'}, 'note': ''}]},
}
SEARCH_ROWS = {
    # 搜索命中的是别名行，名字可能与输入毫无关系（实测搜 'Ono Ryouko' 会命中
    # 另一个人的某个别名行）。这里 s2 就扮演那个捣乱的人。
    'Alpha One': [{'id': 's2', 'aid': 'a8', 'name': 'Alpha Fake', 'original': ''},
                  {'id': 's1', 'aid': 'a1', 'name': 'Alpha One',
                   'original': 'アルファ', 'ismain': True}],
    'One': [{'id': 's1', 'aid': 'a2', 'name': 'A. One', 'original': ''},
            {'id': 's2', 'aid': 'a8', 'name': 'One Fake', 'original': ''}],
    'Alpha': [{'id': 's2', 'aid': 'a8', 'name': 'Alpha Fake', 'original': ''}],
}


def find_sid(node):
    if isinstance(node, list):
        for part in node:
            got = find_sid(part)
            if got:
                return got
    elif isinstance(node, str) and node[:1] == 's' and node[1:].isdigit():
        return node
    return None


def vndb(fail_vn_for=None):
    """按上面的数据回放。fail_vn_for 让某个人的 /vn 翻页请求返回 400。"""
    def handler(path, body, nth):
        filters = body.get('filters') or []
        sid = find_sid(filters)
        head = filters[0] if filters else ''
        if path == 'staff':
            if head == 'or':
                rows = [r for f in filters[1:] for r in STAFF_ROWS.get(f[2], [])]
            elif head == 'search':
                rows = SEARCH_ROWS.get(filters[2], [])
            else:
                rows = STAFF_ROWS.get(sid, [])
            return page(rows, count=len(rows))
        if path == 'character':
            rows = CHAR_ROWS.get(sid, [])
            return page(rows, count=len(rows))
        if path == 'vn':
            if sid == fail_vn_for and body.get('results') == api.PAGE_SIZE:
                return http_error(400, b'boom')
            vids = []
            for char in CHAR_ROWS.get(sid, []):
                for vn in char['vns']:
                    if vn['id'] not in vids:
                        vids.append(vn['id'])
            return page([VN_ROWS[v] for v in sorted(vids)], count=len(vids))
        raise AssertionError('没预料到的 path: %r' % path)
    return handler


# ---------------- 解析 ----------------
def test_load_staff_collects_aliases(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    staff = fetch.load_staff('s1', api.Client())
    assert (staff.name, staff.original) == ('Alpha One', 'アルファ')
    assert staff.alias('a2') == ('A. One', '')
    assert staff.alias('missing') == ('', '')
    assert staff.label() == 'アルファ (Alpha One) s1'


def test_resolve_unknown_id(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    res = fetch.resolve_target('s9', api.Client())
    assert not res.ok and '没有 s9' in res.error


def test_resolve_name_takes_the_only_exact_match(monkeypatch):
    # 搜索同时命中了 s2 的一个别名行，但只有 s1 与输入精确相等。
    FakeApi(vndb()).install(monkeypatch)
    res = fetch.resolve_target('Alpha One', api.Client())
    assert res.ok and res.staff.sid == 's1'


def test_resolve_name_refuses_when_ambiguous(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    res = fetch.resolve_target('One', api.Client())
    assert not res.ok and '命中 2 个人' in res.error
    assert [c.sid for c in res.candidates] == ['s1', 's2']
    # 候选给的是各 id 的主名，不是命中的那行别名（'A. One' / 'One Fake'）
    assert res.candidates[0].label() == 'アルファ (Alpha One) s1'
    assert 'ベータ (Beta Two) s2' in fetch.describe_problem(res)


def test_resolve_name_refuses_single_inexact_hit(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    res = fetch.resolve_target('Alpha', api.Client())
    assert not res.ok and '只命中一个人' in res.error
    assert [c.sid for c in res.candidates] == ['s2']


def test_resolve_name_not_found(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    res = fetch.resolve_target('Nobody Here', api.Client())
    assert not res.ok and '搜不到' in res.error and res.candidates == []


# ---------------- 抓取 ----------------
def test_fetch_credits_drops_other_seiyuu_and_overmatched_vns(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    client = api.Client()
    staff = fetch.load_staff('s1', client)
    seen = []
    credits = fetch.fetch_credits(staff, client, lambda d, t: seen.append((d, t)))
    # v2 是被嵌套 filter 带回来的候选（c2 在里面出现过，但由别人配），整部丢掉；
    # v1 里 s2 配 c3 的那条 va 记录也丢掉。
    assert [(c.vid, c.cid) for c in credits] == [('v1', 'c1'), ('v1', 'c2')]
    assert seen == [(1, 2), (2, 2)]
    one, two = credits
    assert (one.title, one.title_ja, one.released) == ('Game One', 'ゲーム壱', '1995')
    assert (one.cast, one.cast_ja, one.role) == ('Chara One', 'キャラ壱', 'main')
    assert (one.alias, one.alias_ja) == ('Alpha One', 'アルファ')
    assert (two.alias, two.role, two.note) == ('A. One', 'side', '2024 remake')


def test_ensure_resolved_caches_until_cleared(monkeypatch):
    fake = FakeApi(vndb()).install(monkeypatch)
    ctx = RunContext()
    client = api.Client(ctx)
    first = fetch.ensure_resolved({'staff': 's1, s2'}, ctx, client)
    spent = len(fake.calls)
    assert [r.staff.sid for r in first] == ['s1', 's2']
    assert fetch.ensure_resolved({'staff': 's1,s2'}, ctx, client) is first
    assert len(fake.calls) == spent
    fetch.clear_cache(ctx)
    assert fetch.ensure_resolved({'staff': 's1, s2'}, ctx, client) is not first
    assert len(fake.calls) > spent


def test_ensure_counts_is_cached(monkeypatch):
    fake = FakeApi(vndb()).install(monkeypatch)
    ctx = RunContext()
    client = api.Client(ctx)
    staff = fetch.load_staff('s1', client)
    assert fetch.ensure_counts(staff, ctx, client) == (2, 2)
    spent = len(fake.calls)
    assert fetch.ensure_counts(staff, ctx, client) == (2, 2)
    assert len(fake.calls) == spent


def test_ensure_credits_isolates_one_persons_failure(monkeypatch):
    fake = FakeApi(vndb(fail_vn_for='s2')).install(monkeypatch)
    ctx = CountingCtx(10 ** 9)
    client = api.Client(ctx)
    staffs = [fetch.load_staff('s1', client), fetch.load_staff('s2', client)]
    items, failures = fetch.ensure_credits(staffs, ctx, client)
    assert [i.staff.sid for i in items] == ['s1']
    assert len(failures) == 1
    assert failures[0][0].endswith('s2') and 'boom' in failures[0][1]
    # 进度横跨所有人：分母是两人候选数之和
    assert ctx.progress_calls[0] == (1, 4, 'Alpha One 1/2')
    spent = len(fake.calls)
    assert fetch.ensure_credits(staffs, ctx, client) == (items, failures)
    assert len(fake.calls) == spent


def test_cancel_propagates_through_ensure_credits(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    ctx = CountingCtx(10 ** 9)
    client = api.Client(ctx)
    staff = fetch.load_staff('s1', client)
    ctx.after, ctx.checks = 3, 0
    with pytest.raises(Cancelled):
        fetch.ensure_credits([staff], ctx, client)
    assert 'credits' not in ctx.session      # 取消不留半份缓存


# ---------------- 交集 ----------------
def credit(vid, cast, cid, released='2000-01-01'):
    return Credit(vid=vid, title='Game ' + vid, released=released,
                  cid=cid, cast=cast)


def person(sid, credits):
    return StaffCredits(staff=Staff(sid=sid, name=sid.upper()), credits=credits)


def test_intersect_of_three_people():
    a = person('s1', [credit('v1', 'A1', 'c1'), credit('v2', 'A2', 'c2')])
    b = person('s2', [credit('v1', 'B1', 'c3'), credit('v3', 'B3', 'c4')])
    c = person('s3', [credit('v1', 'C1', 'c5'), credit('v2', 'C2', 'c6')])
    common = fetch.intersect([a, b, c])
    assert [x.vid for x in common] == ['v1']
    assert common[0].casts == [[('A1', url_for('c1'))],
                               [('B1', url_for('c3'))],
                               [('C1', url_for('c5'))]]


def test_intersect_keeps_multiple_roles_and_dedupes_same_name():
    a = person('s1', [credit('v1', 'Same', 'c1'), credit('v1', 'Same', 'c9'),
                      credit('v1', 'Other', 'c2')])
    b = person('s2', [credit('v1', 'B', 'c3')])
    common = fetch.intersect([a, b])
    assert common[0].casts[0] == [('Same', url_for('c1')), ('Other', url_for('c2'))]


def test_intersect_sorts_year_only_before_full_date():
    a = person('s1', [credit('v1', 'x', 'c1', released='1995'),
                      credit('v2', 'x', 'c2', released='1995-01-31'),
                      credit('v3', 'x', 'c3', released=None)])
    b = person('s2', [credit(v, 'y', 'c9') for v in ('v1', 'v2', 'v3')])
    assert [x.vid for x in fetch.intersect([a, b])] == ['v1', 'v2', 'v3']


def test_intersect_edge_cases():
    assert fetch.intersect([]) == []
    lone = person('s1', [credit('v1', 'A', 'c1')])
    assert [x.vid for x in fetch.intersect([lone])] == ['v1']
    other = person('s2', [credit('v9', 'B', 'c2')])
    assert fetch.intersect([lone, other]) == []


def test_combos_covers_every_subset_of_three():
    a = person('s1', [credit('v1', 'A1', 'c1'), credit('v2', 'A2', 'c2')])
    b = person('s2', [credit('v1', 'B1', 'c3'), credit('v2', 'B2', 'c4'),
                      credit('v3', 'B3', 'c5')])
    c = person('s3', [credit('v1', 'C1', 'c6'), credit('v3', 'C3', 'c7')])
    got = fetch.combos([a, b, c])
    # 人数降序、输入顺序次之；没有共同作品的组合也在列表里，entries 为空。
    assert [(x.members, [e.vid for e in x.entries]) for x in got] == [
        ((0, 1, 2), ['v1']),
        ((0, 1), ['v1', 'v2']),
        ((0, 2), ['v1']),
        ((1, 2), ['v1', 'v3']),
    ]
    # 每张表只有组合内那几个人的角色列，且与 members 同序。
    pair = got[3]
    assert pair.entries[1].casts == [[('B3', url_for('c5'))],
                                     [('C3', url_for('c7'))]]


def test_combos_matches_intersect_and_is_sorted():
    a = person('s1', [credit('v1', 'x', 'c1', released='1995-01-31'),
                      credit('v2', 'x', 'c2', released='1995')])
    b = person('s2', [credit('v1', 'y', 'c3'), credit('v2', 'y', 'c4')])
    got = fetch.combos([a, b])
    assert len(got) == 1 and got[0].members == (0, 1)
    # 年份粒度排在完整日期之前，和 intersect 同一套排序
    assert [e.vid for e in got[0].entries] == ['v2', 'v1']
    assert [e.vid for e in fetch.intersect([a, b])] == ['v2', 'v1']


def test_combos_of_a_lone_person_is_empty():
    assert fetch.combos([person('s1', [credit('v1', 'A', 'c1')])]) == []
    assert fetch.combos([]) == []


# ---------------- 写盘 ----------------
# Excel 对表名与表结构的要求，openpyxl 一概不校验：违反了要等用户打开文件、看到
# 「已修复的记录: 表」才发现。这里把规则写成断言，套在每个建簿用例的产出上。
CELL_REF_RE = re.compile(r'^(?:[A-Za-z]{1,3}\d{1,7}|[Rr]\d+[Cc]\d+)$')


def check_excel_tables(wb):
    seen = set()
    for ws in wb.worksheets:
        # TableList.items() 给的是 (名字, ref)，Table 对象要从 values() 拿。
        for table in ws.tables.values():
            name = table.displayName
            assert name == table.name
            assert not CELL_REF_RE.match(name), '表名 %r 与单元格引用同形' % name
            assert name.casefold() not in ('c', 'r'), '%r 是保留名' % name
            assert ' ' not in name and (name[0].isalpha() or name[0] in '_\\')
            assert name.casefold() not in seen, '表名 %r 在工作簿内重复' % name
            seen.add(name.casefold())
            rows = ws[table.ref]
            assert len(rows) >= 2, '只有表头的 ref 会让 Excel 判定文件损坏'
            header = [c.value for c in rows[0]]
            assert all(isinstance(h, str) and h for h in header)
            assert len(set(header)) == len(header), '列名不能重复'
            assert [c.name for c in table.tableColumns] == header


def test_table_name_dedupes_case_insensitively():
    used = set()
    assert xlsx.table_name(used, 'Staff_s1') == 'Staff_s1'
    assert xlsx.table_name(used, 'staff_S1') == 'staff_S1_2'


def test_sheet_title_truncates_and_dedupes():
    used = set()
    assert xlsx.sheet_title(used, 'a/b:c') == 'a_b_c'
    long_name = 'x' * 40
    assert xlsx.sheet_title(used, long_name) == 'x' * 31
    assert xlsx.sheet_title(used, long_name) == 'x' * 29 + '_2'
    assert xlsx.sheet_title(used, '') == 'Sheet'


def test_workbook_name():
    one = Staff(sid='s1', name='S1')
    two = Staff(sid='s2', name='S2')
    three = Staff(sid='s3', name='S3')
    assert xlsx.workbook_name([one]) == 'vndb_S1_voiced.xlsx'
    assert xlsx.workbook_name([one, two]) == '共同出演_S1_S2.xlsx'
    assert xlsx.workbook_name([one, two, three]) == '共同出演_S1_S2_S3_3人.xlsx'


def test_workbook_name_writes_every_romaji_in_full():
    staffs = [Staff(sid='s367', name='Ono Ryouko'),
              Staff(sid='s131', name='Mizuhashi Kaori'),
              Staff(sid='s359', name='Okajima Tae')]
    assert xlsx.workbook_name(staffs) == (
        '共同出演_Ono_Ryouko_Mizuhashi_Kaori_Okajima_Tae_3人.xlsx')


def test_workbook_name_falls_back_when_it_would_be_too_long():
    many = [Staff(sid='s%d' % i, name='Nagai Namae %d' % i) for i in range(9)]
    name = xlsx.workbook_name(many)
    assert name == '共同出演_Nagai_Namae_0_Nagai_Namae_1等9人.xlsx'
    assert len(name) <= xlsx.MAX_STEM + len('.xlsx')


def rich(vid, title, cast, cid, released='1995', role='main', note=''):
    return Credit(vid=vid, title=title, title_ja=title + '（日）',
                  released=released, cid=cid, cast=cast, cast_ja=cast + '子',
                  alias='Alias', alias_ja='別名', note=note, role=role)


def test_build_two_people_workbook(tmp_path):
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    a = person('s1', [rich('v1', '=Game One', 'Chara', 'c1'),
                      rich('v2', 'Game Two', 'Other', 'c2', role='side')])
    b = person('s2', [rich('v1', '=Game One', 'Bee', 'c3', role='appears')])
    common = fetch.combos([a, b])
    path = xlsx.save([a, b], common, str(tmp_path))
    assert os.path.basename(path) == '共同出演_S1_S2.xlsx'

    wb = load_workbook(path)
    assert wb.sheetnames == ['概览', '共同出演', 'S1 s1', 'S2 s2']
    check_excel_tables(wb)
    assert [t for ws in wb.worksheets for t in ws.tables] == [
        'Overview', 'Common', 'Staff_s1', 'Staff_s2']
    over = wb['概览']
    assert [c.value for c in over[1]] == list(xlsx.OVERVIEW_HEADERS)
    assert [c.value for c in over[2]] == ['S1', 'S1', 's1', 2, 2, 1, 0, 1, 0, 0]
    assert over['C2'].hyperlink.target == 'https://vndb.org/s1'
    assert over.freeze_panes == 'A2' and len(over.tables) == 1

    both = wb['共同出演']
    assert [c.value for c in both[1]] == ['Released', 'Title', 'Title1', 'S1', 'S2']
    assert [c.value for c in both[2]] == ['1995', '=Game One', '=Game One（日）',
                                          'Chara子', 'Bee子']
    # 以 '=' 开头的标题必须按文本存，否则 Excel 会当公式
    assert both['B2'].data_type == 's'
    assert both['B2'].hyperlink.target == 'https://vndb.org/v1'
    assert both['D2'].hyperlink.target == 'https://vndb.org/c1'

    sheet = wb['S1 s1']
    assert [c.value for c in sheet[1]] == list(xlsx.HEADERS)
    assert [c.value for c in sheet[2]] == ['=Game One', '1995', 'Chara', 'Alias',
                                           None, '=Game One（日）', 'Chara子',
                                           '別名', 'main']
    assert sheet['G2'].hyperlink.target == 'https://vndb.org/c1'
    assert sheet.column_dimensions['A'].width == xlsx.WIDTHS[0]


def test_orig_columns_hold_the_latin_text_when_there_is_no_original(tmp_path):
    """原名列不能因为「原名本就是英文」而空着，链接也要照样加。"""
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    bare = Credit(vid='v1', title='Game One', released='1995', cid='c1',
                  cast='Chara One', alias='Alias One', role='main')
    a = person('s1', [bare])
    b = person('s2', [Credit(vid='v1', title='Game One', released='1995',
                             cid='c3', cast='Bee', role='side')])
    common = fetch.combos([a, b])
    wb = load_workbook(xlsx.save([a, b], common, str(tmp_path)))

    sheet = wb['S1 s1']
    assert [sheet[c + '2'].value for c in 'FGH'] == ['Game One', 'Chara One',
                                                    'Alias One']
    assert sheet['F2'].hyperlink.target == 'https://vndb.org/v1'
    assert sheet['G2'].hyperlink.target == 'https://vndb.org/c1'

    both = wb['共同出演']
    assert [c.value for c in both[2]] == ['1995', 'Game One', 'Game One',
                                          'Chara One', 'Bee']
    assert both['C2'].hyperlink.target == 'https://vndb.org/v1'


def test_same_person_twice_still_gets_unique_table_names(tmp_path):
    """`s367, Ono Ryouko` 会解析出同一个人两次，表名撞名同样会让 Excel 修复。"""
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    a = person('s1', [rich('v1', 'Game One', 'Chara', 'c1')])
    wb = load_workbook(xlsx.save([a, a], fetch.combos([a, a]), str(tmp_path)))
    check_excel_tables(wb)
    assert [t for ws in wb.worksheets for t in ws.tables] == [
        'Overview', 'Common', 'Staff_s1', 'Staff_s1_2']


def test_combo_label_degrades_until_it_fits():
    def staffs(*names):
        return [Staff(sid='s%d' % i, name=n) for i, n in enumerate(names, 1)]

    assert xlsx.combo_label(staffs('Ono Ryouko', 'Mizuhashi Kaori')) == (
        'Ono Ryouko+Mizuhashi Kaori')
    # 三个人的完整罗马音是 39 字符，超了 31，整组一起降到姓 + 名首字母
    assert xlsx.combo_label(
        staffs('Ono Ryouko', 'Mizuhashi Kaori', 'Okajima Tae')) == (
            'Ono R.+Mizuhashi K.+Okajima T.')
    five = xlsx.combo_label(staffs(*['Nagai Namae'] * 5))
    assert five == 'Nagai+Nagai+Nagai+Nagai+Nagai'
    assert len(five) <= xlsx.SHEET_LIMIT
    # 姓也塞不下时退到 sid
    assert xlsx.combo_label(staffs(*['Verylongfamilyname'] * 3)) == 's1+s2+s3'


def trio():
    def cr(vid, cast, cid):
        return Credit(vid=vid, title='Game ' + vid, title_ja='ゲーム' + vid,
                      released='1995', cid=cid, cast=cast, role='main')

    return [StaffCredits(staff=Staff(sid='s367', name='Ono Ryouko'),
                         credits=[cr('v1', 'A1', 'c1'), cr('v2', 'A2', 'c2')]),
            StaffCredits(staff=Staff(sid='s131', name='Mizuhashi Kaori'),
                         credits=[cr('v1', 'B1', 'c3'), cr('v2', 'B2', 'c4'),
                                  cr('v3', 'B3', 'c5')]),
            StaffCredits(staff=Staff(sid='s359', name='Okajima Tae'),
                         credits=[cr('v1', 'C1', 'c6'), cr('v3', 'C3', 'c7')])]


def test_three_people_get_a_sheet_per_combination(tmp_path):
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    items = trio()
    path = xlsx.save(items, fetch.combos(items), str(tmp_path))
    assert os.path.basename(path) == (
        '共同出演_Ono_Ryouko_Mizuhashi_Kaori_Okajima_Tae_3人.xlsx')

    wb = load_workbook(path)
    check_excel_tables(wb)
    assert wb.sheetnames == [
        '概览', '组合',
        'Ono R.+Mizuhashi K.+Okajima T.',
        'Ono Ryouko+Mizuhashi Kaori', 'Ono Ryouko+Okajima Tae',
        'Mizuhashi Kaori+Okajima Tae',
        'Ono Ryouko s367', 'Mizuhashi Kaori s131', 'Okajima Tae s359']
    assert [t for ws in wb.worksheets for t in ws.tables] == [
        'Overview', 'Combos', 'Common_s367_s131_s359', 'Common_s367_s131',
        'Common_s367_s359', 'Common_s131_s359',
        'Staff_s367', 'Staff_s131', 'Staff_s359']

    # 每张组合表只有组合内那几个人的列，且与组合同序
    pair = wb['Mizuhashi Kaori+Okajima Tae']
    assert [c.value for c in pair[1]] == ['Released', 'Title', 'Title1',
                                          'Mizuhashi Kaori', 'Okajima Tae']
    assert [r[1] for r in pair.iter_rows(min_row=2, values_only=True)] == [
        'Game v1', 'Game v3']
    assert [r[4] for r in pair.iter_rows(min_row=2, values_only=True)] == [
        'C1', 'C3']


def test_combo_index_maps_full_names_to_sheets(tmp_path):
    """页名被 31 字符截过，完整罗马音只能靠这一页查，还要能点着跳过去。"""
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    items = trio()
    wb = load_workbook(xlsx.save(items, fetch.combos(items), str(tmp_path)))
    index = wb['组合']
    assert [c.value for c in index[1]] == list(xlsx.COMBO_HEADERS)
    assert [tuple(r) for r in index.iter_rows(min_row=2, values_only=True)] == [
        (3, 'Ono Ryouko、Mizuhashi Kaori、Okajima Tae', 1,
         'Ono R.+Mizuhashi K.+Okajima T.'),
        (2, 'Ono Ryouko、Mizuhashi Kaori', 2, 'Ono Ryouko+Mizuhashi Kaori'),
        (2, 'Ono Ryouko、Okajima Tae', 1, 'Ono Ryouko+Okajima Tae'),
        (2, 'Mizuhashi Kaori、Okajima Tae', 2, 'Mizuhashi Kaori+Okajima Tae')]
    # 内部跳转要写 location；写成 target 会被当成外部 URL
    assert index['D2'].hyperlink.location == (
        "'Ono R.+Mizuhashi K.+Okajima T.'!A1")


def test_combination_without_shared_works_gets_no_sheet(tmp_path):
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    a = person('s1', [credit('v1', 'A', 'c1')])
    b = person('s2', [credit('v1', 'B', 'c2')])
    c = person('s3', [credit('v9', 'C', 'c3')])
    wb = load_workbook(xlsx.save([a, b, c], fetch.combos([a, b, c]),
                                 str(tmp_path)))
    check_excel_tables(wb)
    assert wb.sheetnames == ['概览', '组合', 'S1+S2', 'S1 s1', 'S2 s2', 'S3 s3']
    index = wb['组合']
    assert [tuple(r) for r in index.iter_rows(min_row=2, values_only=True)] == [
        (3, 'S1、S2、S3', 0, '（无共同出演，未建表）'),
        (2, 'S1、S2', 1, 'S1+S2'),
        (2, 'S1、S3', 0, '（无共同出演，未建表）'),
        (2, 'S2、S3', 0, '（无共同出演，未建表）')]
    assert index['D2'].hyperlink is None


def test_build_single_person_has_no_common_sheet(tmp_path):
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    lone = person('s1', [])
    path = xlsx.save([lone], [], str(tmp_path))
    assert os.path.basename(path) == 'vndb_S1_voiced.xlsx'
    wb = load_workbook(path)
    assert wb.sheetnames == ['概览', 'S1 s1']
    # 空表不能套 Table：只有表头的 ref 会让 Excel 判定文件损坏
    assert len(wb['S1 s1'].tables) == 0
    assert wb['S1 s1'].freeze_panes == 'A2'

    again = xlsx.save([lone], [], str(tmp_path))
    assert os.path.basename(again) == 'vndb_S1_voiced_1.xlsx'


def test_save_accepts_an_explicit_xlsx_path(tmp_path):
    pytest.importorskip('openpyxl')
    target = tmp_path / 'sub' / '我的.xlsx'
    path = xlsx.save([person('s1', [])], [], str(target))
    assert path == str(target) and os.path.exists(path)


def test_clean_strips_control_chars():
    assert xlsx.clean('a\x01b\nc') == 'ab\nc'
    assert xlsx.clean(5) == 5


# ---------------- 工具层 ----------------
def test_validate_is_offline_and_reports_first_bad_target(tmp_path):
    assert tool.validate({'staff': 's1', 'out_dir': str(tmp_path)}) == []
    errors = dict(tool.validate({'staff': 'v3, c9',
                                 'out_dir': str(tmp_path / '没有这个目录')}))
    assert '目录不存在' in errors['out_dir']
    assert '作品' in errors['staff']          # 只报第一个坏目标
    assert dict(tool.validate({'staff': '、、'}))['staff'] == '没解析出任何目标'


def test_validate_caps_the_number_of_people():
    ok = ', '.join('s%d' % i for i in range(1, tool.MAX_TARGETS + 1))
    assert tool.validate({'staff': ok}) == []
    too_many = ok + ', s99'
    message = dict(tool.validate({'staff': too_many}))['staff']
    assert '最多 %d 个人' % tool.MAX_TARGETS in message
    assert '502 个两两及以上的组合' in message      # 2^9 - 9 - 1


def test_eta_text_switches_unit():
    assert tool.eta_text(0, 1) == '约 5 秒'
    assert tool.eta_text(2000, 3) == '约 2 分钟'


def test_refresh_clears_cache_only_once_per_toggle():
    ctx = RunContext()
    ctx.session['counts'] = {'s1': (1, 1)}
    tool._apply_refresh({'refresh': True}, ctx)
    assert 'counts' not in ctx.session

    ctx.session['counts'] = {'s1': (1, 1)}
    tool._apply_refresh({'refresh': True}, ctx)
    assert 'counts' in ctx.session           # 同一次勾选不再清第二遍

    tool._apply_refresh({'refresh': False}, ctx)
    tool._apply_refresh({'refresh': True}, ctx)
    assert 'counts' not in ctx.session       # 取掉再勾上又能清


def test_preview_lists_each_person(monkeypatch, tmp_path):
    FakeApi(vndb()).install(monkeypatch)
    result = tool.preview({'staff': 's1, s2', 'out_dir': str(tmp_path)},
                          RunContext())
    assert result.ok
    assert 'アルファ (Alpha One) s1：角色 2，作品 ≤2' in result.summary
    assert '共 2 人，预计' in result.summary
    assert '共同出演_Alpha_One_Beta_Two.xlsx' in result.summary


def test_preview_without_target_asks_for_one():
    result = tool.preview({'staff': ''}, RunContext())
    assert not result.ok and 's367' in result.summary


def test_preview_refuses_ambiguous_name(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    result = tool.preview({'staff': 'One'}, RunContext())
    assert not result.ok
    assert '命中 2 个人' in result.summary
    assert 'Alpha One' in result.summary and 'Beta Two' in result.summary


def test_preview_warns_about_the_lonely_single_person(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    result = tool.preview({'staff': 's1'}, RunContext())
    assert result.ok and result.warnings == ['只有一个人，不会有共同出演页。']


def test_preview_reports_the_combinations_for_three_people(monkeypatch, tmp_path):
    FakeApi(vndb()).install(monkeypatch)
    result = tool.preview({'staff': 's1, s2, s3', 'out_dir': str(tmp_path)},
                          RunContext())
    assert result.ok
    assert '共同出演_Alpha_One_Beta_Two_Gamma_Three_3人.xlsx' in result.summary
    assert '共同出演表：4 个组合' in result.summary
    assert result.warnings == []          # 三个人还不至于要警告


def test_run_writes_a_workbook_and_reuses_the_preview_cache(monkeypatch, tmp_path):
    pytest.importorskip('openpyxl')
    fake = FakeApi(vndb()).install(monkeypatch)
    ctx = RunContext()
    params = {'staff': 's1, s2', 'out_dir': str(tmp_path)}
    tool.preview(params, ctx)
    before = len(fake.calls)             # 解析 2 + count 4
    result = tool.run(params, ctx)

    # run 只补抓取本身的 /character 与 /vn，解析与 count 全部命中缓存。
    assert len(fake.calls) - before == 4
    assert fake.paths()[before:] == ['character', 'vn', 'character', 'vn']
    assert result.failures == []
    assert result.warnings == []
    assert '共同出演 : 1 部' in result.summary
    assert 'アルファ (Alpha One) s1 : 1 部作品 / 2 个角色' in result.summary
    path = result.output_paths[0]
    assert os.path.basename(path) == '共同出演_Alpha_One_Beta_Two.xlsx'
    assert os.path.exists(path)


def test_run_skips_targets_it_cannot_resolve(monkeypatch, tmp_path):
    pytest.importorskip('openpyxl')
    FakeApi(vndb()).install(monkeypatch)
    ctx = CountingCtx(after=10 ** 6)
    result = tool.run({'staff': 's1, s9', 'out_dir': str(tmp_path)}, ctx)
    assert result.failures == [('s9', 'vndb 上没有 s9 这个人')]
    assert result.warnings == ['1 个目标没定位到人，已跳过。']
    assert os.path.exists(result.output_paths[0])
    assert any('s9' in msg for level, msg in ctx.logs if level == 'warn')


def test_run_without_anyone_writes_nothing(monkeypatch, tmp_path):
    FakeApi(vndb()).install(monkeypatch)
    result = tool.run({'staff': 's9', 'out_dir': str(tmp_path)}, RunContext())
    assert result.output_paths == []
    assert len(result.failures) == 1
    assert list(tmp_path.iterdir()) == []


def test_run_reports_a_persons_fetch_failure(monkeypatch, tmp_path):
    pytest.importorskip('openpyxl')
    FakeApi(vndb(fail_vn_for='s2')).install(monkeypatch)
    result = tool.run({'staff': 's1, s2', 'out_dir': str(tmp_path)}, RunContext())
    assert [name for name, _ in result.failures] == ['ベータ (Beta Two) s2']
    assert result.warnings == ['1 个人抓取失败，已跳过。']
    # 只剩一个人了，就不该再有共同出演页。
    assert '共同出演' not in result.summary
    assert os.path.basename(result.output_paths[0]) == 'vndb_Alpha_One_voiced.xlsx'


def test_run_cancelled_carries_an_empty_partial(monkeypatch, tmp_path):
    FakeApi(vndb()).install(monkeypatch)

    def boom(*args, **kwargs):
        raise Cancelled()

    monkeypatch.setattr(fetch, 'ensure_credits', boom)
    with pytest.raises(Cancelled) as caught:
        tool.run({'staff': 's1', 'out_dir': str(tmp_path)}, RunContext())
    assert '取消' in caught.value.partial.summary
    assert list(tmp_path.iterdir()) == []


def test_tool_spec_fields_match_what_run_reads():
    keys = {f.key for f in tool.TOOL.fields}
    assert keys == {'staff', 'out_dir', 'refresh'}
    assert [f.key for f in tool.TOOL.fields if f.rescan] == ['staff', 'refresh']


# ---------------- 结果表格 ----------------
def test_preview_table_lists_the_people_with_clickable_ids(monkeypatch, tmp_path):
    FakeApi(vndb()).install(monkeypatch)
    table = tool.preview({'staff': 's1, s2', 'out_dir': str(tmp_path)},
                         RunContext()).table
    assert table.columns == ('声优', '罗马字', 'ID', '角色数', '作品数 ≤')
    assert table.rows[0] == ('アルファ', 'Alpha One',
                             ('s1', 'https://vndb.org/s1'), 2, 2)
    assert isinstance(table.rows[0][3], int)     # 数字列要按数值排，不能是字符串


def test_preview_table_switches_to_candidates_when_ambiguous(monkeypatch):
    FakeApi(vndb()).install(monkeypatch)
    table = tool.preview({'staff': 'One'}, RunContext()).table
    assert table.columns[0] == '填的是'
    assert [r[0] for r in table.rows] == ['One', 'One']
    assert [r[3] for r in table.rows] == [('s1', 'https://vndb.org/s1'),
                                          ('s2', 'https://vndb.org/s2')]


def test_preview_without_target_has_no_table():
    assert tool.preview({'staff': ''}, RunContext()).table is None


def test_run_table_is_the_common_works(monkeypatch, tmp_path):
    pytest.importorskip('openpyxl')
    FakeApi(vndb()).install(monkeypatch)
    table = tool.run({'staff': 's1, s2', 'out_dir': str(tmp_path)},
                     RunContext()).table
    assert table.columns == ('发售日', 'Title', '日文原名', 'Alpha One', 'Beta Two')
    assert table.title == '共同出演 1 部'
    assert table.rows == [('1995', ('Game One', 'https://vndb.org/v1'),
                           ('ゲーム壱', 'https://vndb.org/v1'),
                           # 一个人在同一部里配两个角色，链接指向谁都不对
                           ('キャラ壱 / Chara Two', None),
                           ('キャラ参', 'https://vndb.org/c3'))]


def test_common_table_falls_back_when_a_title_has_no_original():
    """屏幕上的表和文件里的表要一致：没有日文原名就摆英文标题，不留空格。"""
    items = [person('s1', []), person('s2', [])]
    common = [Common(vid='v1', title='Game One', released='1995',
                     casts=[[('Chara', url_for('c1'))], [('Bee', url_for('c3'))]])]
    row = tool._common_table(items, common).rows[0]
    assert row[2] == ('Game One', 'https://vndb.org/v1')


def test_run_table_falls_back_to_one_persons_credits(monkeypatch, tmp_path):
    pytest.importorskip('openpyxl')
    FakeApi(vndb()).install(monkeypatch)
    table = tool.run({'staff': 's1', 'out_dir': str(tmp_path)},
                     RunContext()).table
    assert table.columns == ('发售日', 'Title', '角色', 'As', 'Role')
    assert table.title == 'アルファ (Alpha One) s1：2 条出演记录'
    assert [r[2] for r in table.rows] == [('キャラ壱', 'https://vndb.org/c1'),
                                          ('Chara Two', 'https://vndb.org/c2')]
    assert [r[4] for r in table.rows] == ['main', 'side']


def test_run_writes_every_combination_for_three_people(monkeypatch, tmp_path):
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    FakeApi(vndb()).install(monkeypatch)
    result = tool.run({'staff': 's1, s2, s3', 'out_dir': str(tmp_path)},
                      RunContext())
    assert os.path.basename(result.output_paths[0]) == (
        '共同出演_Alpha_One_Beta_Two_Gamma_Three_3人.xlsx')
    assert '共同出演 : 4 个组合有交集（共 4 个组合）' in result.summary
    assert '  Alpha One、Beta Two : 1 部' in result.summary
    assert '  Beta Two、Gamma Three : 2 部' in result.summary

    wb = load_workbook(result.output_paths[0])
    check_excel_tables(wb)
    assert wb.sheetnames[:2] == ['概览', '组合']
    assert len([n for n in wb.sheetnames if '+' in n]) == 4

    # 屏幕上摆全员那一档，列头是三个人
    assert result.table.columns == ('发售日', 'Title', '日文原名',
                                    'Alpha One', 'Beta Two', 'Gamma Three')
    assert result.table.title == '共同出演 1 部'


def test_same_person_named_twice_is_merged(monkeypatch, tmp_path):
    """`s1, Alpha One` 指向同一个人：抓一遍、一页明细，不出「他和他自己」那张表。"""
    pytest.importorskip('openpyxl')
    from openpyxl import load_workbook

    fake = FakeApi(vndb()).install(monkeypatch)
    result = tool.run({'staff': 's1, Alpha One', 'out_dir': str(tmp_path)},
                      RunContext())
    assert os.path.basename(result.output_paths[0]) == 'vndb_Alpha_One_voiced.xlsx'
    wb = load_workbook(result.output_paths[0])
    assert wb.sheetnames == ['概览', 'アルファ s1']
    assert '1 个目标指向同一个人，已合并。' in result.warnings

    # 抓取次数与只填一次时一致，不再白打一倍
    once = FakeApi(vndb()).install(monkeypatch)
    tool.run({'staff': 's1', 'out_dir': str(tmp_path)}, RunContext())
    resolve_calls = len([c for c in fake.calls if c[0] == 'staff'])
    assert len([c for c in fake.calls if c[0] != 'staff']) == \
        len([c for c in once.calls if c[0] != 'staff'])
    assert resolve_calls > len([c for c in once.calls if c[0] == 'staff'])


def test_preview_merges_the_duplicate_and_says_so(monkeypatch, tmp_path):
    FakeApi(vndb()).install(monkeypatch)
    result = tool.preview({'staff': 's1, Alpha One', 'out_dir': str(tmp_path)},
                          RunContext())
    assert result.ok
    assert result.summary.splitlines() == [
        'アルファ (Alpha One) s1：角色 2，作品 ≤2',
        '共 1 人，预计 约 5 秒；输出 vndb_Alpha_One_voiced.xlsx']
    assert '有 1 个目标指向同一个人，已合并。' in result.warnings
    assert len(result.table.rows) == 1


def test_api_errors_are_shown_as_a_sentence_not_a_traceback(monkeypatch, tmp_path):
    """ApiError 是「已翻译成人话的失败」，穿到 worker 那层会变成一段红色的栈。"""
    FakeApi(lambda path, body, nth: http_error(500, b'down')).install(monkeypatch)
    preview = tool.preview({'staff': 's1', 'out_dir': str(tmp_path)}, RunContext())
    assert not preview.ok
    assert preview.summary.startswith('vndb 接口出错：')
    assert 'HTTP 500' in preview.summary

    result = tool.run({'staff': 's1', 'out_dir': str(tmp_path)}, RunContext())
    assert result.output_paths == []
    assert 'vndb 接口出错' in result.summary
    assert [n for n, _ in result.failures] == ['vndb 接口']


def test_result_table_falls_back_to_the_first_non_empty_combo(monkeypatch,
                                                              tmp_path):
    """三人以上全员没有交集时，原先固定摆全员那一档，屏幕上一片空白。"""
    pytest.importorskip('openpyxl')
    staffs = [Staff(sid='s%d' % i, name='N%d S%d' % (i, i)) for i in (1, 2, 3)]
    items = [person('s1', [credit('v1', 'a', 'c1')]),
             person('s2', [credit('v1', 'b', 'c2'), credit('v2', 'b', 'c3')]),
             person('s3', [credit('v2', 'c', 'c4')])]
    for item, staff in zip(items, staffs):
        item.staff = staff
    monkeypatch.setattr(fetch, 'ensure_resolved',
                        lambda params, ctx, client=None:
                        [Resolution(target=s.sid, staff=s) for s in staffs])
    monkeypatch.setattr(fetch, 'ensure_credits',
                        lambda ss, ctx, client=None: (items, []))

    result = tool.run({'staff': 's1, s2, s3', 'out_dir': str(tmp_path)},
                      RunContext())
    assert '共同出演 : 2 个组合有交集（共 4 个组合）' in result.summary
    assert result.table.title == 'N1 S1、N2 S2 共同出演 1 部'
    assert result.table.columns == ('发售日', 'Title', '日文原名',
                                    'N1 S1', 'N2 S2')
    assert len(result.table.rows) == 1


# ---------------- 命令行 ----------------
def test_cli_check_targets():
    assert cli.check_targets('s1, s2')[0] == ['s1', 's2']
    assert cli.check_targets('')[1] == '没解析出任何目标。'
    assert '作品' in cli.check_targets('https://vndb.org/v3')[1]


def test_cli_shares_the_person_cap_with_the_gui():
    """规则只有一份：以前人数上限只装在 validate 里，命令行完全不受约束。"""
    ok = ', '.join('s%d' % i for i in range(1, tool.MAX_TARGETS + 1))
    assert cli.check_targets(ok) == (fetch.parse_targets(ok), '')
    targets, error = cli.check_targets(ok + ', s99')
    assert targets == [] and '最多 %d 个人' % tool.MAX_TARGETS in error
