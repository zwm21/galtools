# -*- coding: utf-8 -*-
"""写 Excel：一个工作簿 = 概览 + 共同出演 + 每人一页。

openpyxl 一律在函数内 import。tests/test_registry.py 断言 discover() 的 errors
为空，模块级 import 第三方库会让没装 openpyxl 的环境连工具列表都列不出来。

表格用 openpyxl 的原生 Table（自带筛选下拉与隔行底色）而不是手动画格式：几百行
的表在 Excel 里能直接按列排序，可读性差别很大。空表不能加 Table——Excel 要求
ref 至少含一行数据，只有表头的 ref 会让文件被判定为损坏。
"""
import os
import re
from .model import ROLES, url_for

# 每人一页的列。前 8 列沿用旧脚本的表头与列宽（Title1/Cast1/As1 是原名，原名本就
# 是拉丁字母时退回罗马字，不留空列），末尾新增 Role——角色主次是判断「这人在这部里
# 是主役还是路人」的关键信息。
HEADERS = ('Title', 'Released', 'Cast', 'As', 'Note',
           'Title1', 'Cast1', 'As1', 'Role')
WIDTHS = (46, 11, 30, 20, 10, 46, 30, 20, 10)

OVERVIEW_HEADERS = (('声优', '罗马字', 'ID', '作品数', '角色数')
                    + ROLES + ('未标注',))
OVERVIEW_WIDTHS = (26, 22, 8, 8, 8, 8, 9, 8, 9, 9)

COMMON_HEADERS = ('Released', 'Title', 'Title1')
COMMON_WIDTHS = (11, 46, 46)
CAST_WIDTH = 30

# 三人以上时的组合索引页。工作表名只有 31 个字符，放不下三个人的完整罗马音，
# 这一页负责把完整名字与「哪张表」的对应关系摆出来，并提供跳转。
COMBO_HEADERS = ('人数', '声优', '共同作品数', '工作表')
COMBO_WIDTHS = (6, 56, 12, 32)

TABLE_STYLE = 'TableStyleMedium2'
LINK_COLOR = '0563C1'

# Excel 不接受的工作表名字符，以及 xml 里非法的控制字符。
BAD_SHEET_CHARS = re.compile(r'[\[\]:*?/\\]')
CONTROL_CHARS = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
NON_WORD = re.compile(r'[^\w]+', re.UNICODE)

SHEET_LIMIT = 31         # Excel 的工作表名长度上限
# 文件名不设 255 而是 120：出输出目录本身也占路径，整条路径过 260 在 Windows 上
# 会被一堆 API 拒绝，还要给 unique_path 的 _1 序号留位置。
MAX_STEM = 120
OVERVIEW_SHEET = '概览'
COMMON_SHEET = '共同出演'
COMBO_SHEET = '组合'
COMBO_SEP = '+'

# 表名不能与单元格引用同形：'T1' 就是 T 列第 1 行，Excel 会判定文件损坏，打开时
# 报「已修复的记录: 表」并把表重命名。下面这些要么含下划线，要么长过三个字母
# （列名最多到 XFD），因此不可能被解析成引用——'Tbl1' 这种就还是会中招。
OVERVIEW_TABLE = 'Overview'
COMMON_TABLE = 'Common'
COMBO_TABLE = 'Combos'
STAFF_TABLE = 'Staff_%s'


# ---------------- 命名 ----------------
def slug(text):
    return NON_WORD.sub('_', text or '').strip('_') or 'unknown'


def workbook_name(staffs):
    """单人沿用旧脚本的 vndb_<Name>_voiced.xlsx，两人叫 共同出演_A_B.xlsx，
    三人以上把每个人的罗马音都写进去、末尾缀人数。

    名字太多会把路径顶到 Windows 的上限（文件名 255、整条路径 260），所以留了
    退路：拼出来太长就退回前两人 + 等N人。
    """
    names = [slug(s.name or s.sid) for s in staffs]
    if not names:
        return 'vndb_voiced.xlsx'
    if len(names) == 1:
        return 'vndb_%s_voiced.xlsx' % names[0]
    if len(names) == 2:
        return '共同出演_%s_%s.xlsx' % (names[0], names[1])
    stem = '共同出演_%s_%d人' % ('_'.join(names), len(names))
    if len(stem) > MAX_STEM:
        stem = '共同出演_%s_%s等%d人' % (names[0], names[1], len(names))
    return stem + '.xlsx'


def unique_path(path):
    """已存在时追加序号，与 audio_filter 的 unique_dest 同款：x.xlsx -> x_1.xlsx。

    刻意不覆盖：抓一次要十几秒到几分钟，静默覆盖上一次的结果代价太大。
    """
    if not os.path.exists(path):
        return path
    stem, ext = os.path.splitext(path)
    i = 1
    while True:
        cand = '%s_%d%s' % (stem, i, ext)
        if not os.path.exists(cand):
            return cand
        i += 1


def sheet_title(used, base):
    """截到 31 字符、去掉非法字符、按需追加序号。used 存已用过的小写名。"""
    name = BAD_SHEET_CHARS.sub('_', (base or '').strip())[:SHEET_LIMIT].strip()
    name = name or 'Sheet'
    if name.casefold() not in used:
        used.add(name.casefold())
        return name
    i = 2
    while True:
        suffix = '_%d' % i
        cand = name[:SHEET_LIMIT - len(suffix)] + suffix
        if cand.casefold() not in used:
            used.add(cand.casefold())
            return cand
        i += 1


def _name_forms(staff):
    """一个人的名字，从最全到最省：完整罗马音 → 姓 + 名首字母 → 姓 → sid。"""
    name = (staff.name or '').strip()
    parts = name.split()
    forms = [name]
    if len(parts) > 1:
        forms.append('%s %s.' % (parts[0], parts[1][:1]))
        forms.append(parts[0])
    forms.append(staff.sid)
    out = []
    for form in forms:
        if form and form not in out:
            out.append(form)
    return out or ['unknown']


def combo_label(staffs):
    """组合的工作表名。

    Excel 只给 31 个字符，三个人的完整罗马音一定超（`Ono Ryouko+Mizuhashi
    Kaori+Okajima Tae` 是 39），所以整组一起降级，取第一个塞得下的写法。连 sid
    都塞不下时照样返回，交给 sheet_title 截断去重——完整名字在组合索引页里。
    """
    ladders = [_name_forms(s) for s in staffs]
    label = ''
    for level in range(max(len(l) for l in ladders)):
        label = COMBO_SEP.join(l[min(level, len(l) - 1)] for l in ladders)
        if len(label) <= SHEET_LIMIT:
            break
    return label


def table_name(used, base):
    """表名在工作簿内必须唯一，且不区分大小写。

    同一个人填两次（`s367, Ono Ryouko`）就会走到重名这一支——重名同样会让
    Excel 判定文件损坏，和工作表重名不是一回事，得各自去重。
    """
    name = base
    i = 2
    while name.casefold() in used:
        name = '%s_%d' % (base, i)
        i += 1
    used.add(name.casefold())
    return name


# ---------------- 单元格 ----------------
def clean(value):
    """去掉 xml 不接受的控制字符（note 里偶有），非字符串原样返回。"""
    if isinstance(value, str):
        return CONTROL_CHARS.sub('', value)
    return value


def _put(ws, row, col, value, url=None, font=None):
    """写一格。以 '=' 开头的字符串强制按文本存，否则 Excel 会当成公式。"""
    cell = ws.cell(row=row, column=col)
    value = clean(value)
    cell.value = value
    if isinstance(value, str) and value.startswith('='):
        cell.data_type = 's'
    if url and value:
        cell.hyperlink = url
        if font is not None:
            cell.font = font
    return cell


def _row(ws, row, values):
    for col, value in enumerate(values, 1):
        _put(ws, row, col, value)


def _finish(ws, widths, rows, tname):
    """列宽 + 冻结表头 + 套原生表格。rows 为 0 时不加表格。"""
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    for col, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A2'
    if not rows:
        return
    ref = 'A1:%s%d' % (get_column_letter(len(widths)), rows + 1)
    table = Table(displayName=tname, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name=TABLE_STYLE, showRowStripes=True)
    ws.add_table(table)


# ---------------- 各页 ----------------
def _overview(wb, items, title, tname, font):
    ws = wb.create_sheet(title)
    _row(ws, 1, OVERVIEW_HEADERS)
    for row, item in enumerate(items, 2):
        staff = item.staff
        counts = dict.fromkeys(ROLES, 0)
        other = 0
        for credit in item.credits:
            if credit.role in counts:
                counts[credit.role] += 1
            else:
                other += 1
        _put(ws, row, 1, staff.original or staff.name)
        _put(ws, row, 2, staff.name)
        _put(ws, row, 3, staff.sid, url_for(staff.sid), font)
        _put(ws, row, 4, len(item.vids))
        _put(ws, row, 5, len(item.credits))
        for i, role in enumerate(ROLES):
            _put(ws, row, 6 + i, counts[role])
        _put(ws, row, 6 + len(ROLES), other)
    _finish(ws, OVERVIEW_WIDTHS, len(items), tname)


def staff_sheet_base(staff):
    """每人一页的页名：日文原名 + sid，认人比罗马字快。"""
    return '%s %s' % (staff.original or staff.name, staff.sid)


def _staff_sheet(wb, item, title, tname, font):
    ws = wb.create_sheet(title)
    _row(ws, 1, HEADERS)
    for row, credit in enumerate(item.credits, 2):
        vn_url, char_url = url_for(credit.vid), url_for(credit.cid)
        _put(ws, row, 1, credit.title, vn_url, font)
        _put(ws, row, 2, credit.released)
        _put(ws, row, 3, credit.cast, char_url, font)
        _put(ws, row, 4, credit.alias)
        _put(ws, row, 5, credit.note)
        _put(ws, row, 6, credit.title_orig, vn_url, font)
        _put(ws, row, 7, credit.cast_orig, char_url, font)
        _put(ws, row, 8, credit.alias_orig)
        _put(ws, row, 9, credit.role)
    _finish(ws, WIDTHS, len(item.credits), tname)


def cast_columns(items):
    """共同出演页的人名列头。同名时补 sid——Table 不允许重复表头。"""
    names, seen = [], set()
    for item in items:
        name = item.staff.name or item.staff.sid
        if name in seen:
            name = '%s %s' % (name, item.staff.sid)
        seen.add(name)
        names.append(name)
    return names


def _common_sheet(wb, items, combo, title, tname, font):
    """一个组合的共同出演页。列头只含这个组合里的人，与 combo.members 同序。"""
    members = [items[i] for i in combo.members]
    ws = wb.create_sheet(title)
    _row(ws, 1, tuple(COMMON_HEADERS) + tuple(cast_columns(members)))
    for row, entry in enumerate(combo.entries, 2):
        vn_url = url_for(entry.vid)
        _put(ws, row, 1, entry.released)
        _put(ws, row, 2, entry.title, vn_url, font)
        _put(ws, row, 3, entry.title_orig, vn_url, font)
        for i, casts in enumerate(entry.casts):
            # 一个人在同一部里配多个角色时，链接指向谁都不对，索性不加。
            link = casts[0][1] if len(casts) == 1 else None
            _put(ws, row, 4 + i, ' / '.join(t for t, _ in casts), link, font)
    _finish(ws, tuple(COMMON_WIDTHS) + (CAST_WIDTH,) * len(members),
            len(combo.entries), tname)


def _combo_index(wb, items, planned, title, tname, font):
    """组合索引页：完整罗马音 → 哪张表，附跳转链接。

    工作表名被 31 字符截过、还可能因重名加了序号，完整名字必须有地方可查。
    """
    from openpyxl.worksheet.hyperlink import Hyperlink

    ws = wb.create_sheet(title)
    _row(ws, 1, COMBO_HEADERS)
    for row, (combo, sheet) in enumerate(planned, 2):
        _put(ws, row, 1, len(combo.members))
        _put(ws, row, 2, '、'.join(items[i].staff.name or items[i].staff.sid
                                   for i in combo.members))
        _put(ws, row, 3, len(combo.entries))
        cell = _put(ws, row, 4, sheet or '（无共同出演，未建表）')
        if sheet:
            # 内部跳转要写 location 而不是 target，后者会被当成外部 URL。
            cell.hyperlink = Hyperlink(ref=cell.coordinate,
                                       location="'%s'!A1"
                                                % sheet.replace("'", "''"))
            cell.font = font
    _finish(ws, COMBO_WIDTHS, len(planned), tname)


# ---------------- 入口 ----------------
def target_dir(target):
    """save 会把工作簿写进哪个目录。

    工具层在**抓取之前**拿它先建一次目录：抓完几分钟才发现盘不存在，那几分钟
    就白等了。算目录的规则因此只有这一份，否则先挡下的地方和真正写入的地方
    可能不是同一个。
    """
    target = os.path.abspath(target or '.')
    if target.lower().endswith('.xlsx'):
        return os.path.dirname(target) or '.'
    return target


def build(items, combos, path):
    """写出工作簿，返回 path。

    combos 是 fetch.combos 的结果（两两及以上的全部组合，含空组合）。两人时沿用
    旧脚本的单页「共同出演」，哪怕没有交集也留着那张空页；三人以上给每个非空组合
    各一页，并在前面加一张组合索引页。页名与表名都要先定好再写：索引页排在组合页
    之前，但它要引用那些页的最终名字。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    font = Font(color=LINK_COLOR, underline='single')
    wb = Workbook()
    wb.remove(wb.active)
    used, names = set(), set()
    many = len(items) > 2

    overview = sheet_title(used, OVERVIEW_SHEET)
    index = sheet_title(used, COMBO_SHEET) if many else ''
    planned = []
    for combo in combos if len(items) > 1 else ():
        if not many:
            planned.append((combo, sheet_title(used, COMMON_SHEET)))
        elif combo.entries:
            planned.append((combo, sheet_title(
                used, combo_label([items[i].staff for i in combo.members]))))
        else:
            planned.append((combo, ''))

    _overview(wb, items, overview, table_name(names, OVERVIEW_TABLE), font)
    if index:
        _combo_index(wb, items, planned, index,
                     table_name(names, COMBO_TABLE), font)
    for combo, title in planned:
        if not title:
            continue
        base = COMMON_TABLE if not many else 'Common_' + '_'.join(
            items[i].staff.sid or str(i) for i in combo.members)
        _common_sheet(wb, items, combo, title, table_name(names, base), font)
    for item in items:
        _staff_sheet(wb, item, sheet_title(used, staff_sheet_base(item.staff)),
                     table_name(names, STAFF_TABLE % item.staff.sid), font)
    wb.save(path)
    return path


def save(items, combos, target):
    """target 可以是目录也可以是 .xlsx 路径。返回实际写入的路径。"""
    directory = target_dir(target)
    if (target or '').lower().endswith('.xlsx'):
        name = os.path.basename(target)
    else:
        name = workbook_name([i.staff for i in items])
    os.makedirs(directory, exist_ok=True)
    return build(items, combos, unique_path(os.path.join(directory, name)))
